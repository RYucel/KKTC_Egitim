#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
07_excel.py — Üretilen JSON verilerinden konsolide, sekmeli ve formüllü Excel
çalışma kitabını oluşturur (cikti/KGS_Ingilizce_Analiz.xlsx).

Türetilen değerler (yüzde, kohort-arındırma, ortalama, korelasyon) canlı Excel
formülüdür; Excel/LibreOffice'te açıldığında otomatik hesaplanır.
01-04 betikleri önce çalıştırılmış olmalıdır.
"""
import json, csv
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
import ortak

NAVY, BLUE, LBLUE, GREY = "1F3864", "2E5496", "D6E0F0", "808080"
FONT = "Arial"
thin = Side(style="thin", color="BFBFBF"); BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
def hdr(c): c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10); c.fill = PatternFill("solid", fgColor=BLUE); c.alignment = Alignment(horizontal="center", wrap_text=True)
def title(c): c.font = Font(name=FONT, bold=True, color="FFFFFF", size=13); c.fill = PatternFill("solid", fgColor=NAVY)
def cel(c, fmt=None, bold=False, align="center"):
    c.font = Font(name=FONT, size=10, bold=bold); c.alignment = Alignment(horizontal=align)
    if fmt: c.number_format = fmt

def main():
    M = json.load(open(ortak.VERI / "metrikler.json", encoding="utf-8"))
    VOC = {int(k): v for k, v in json.load(open(ortak.VERI / "kelime_cesitliligi.json", encoding="utf-8")).items()}
    RD = {int(k): v for k, v in json.load(open(ortak.VERI / "okunabilirlik.json", encoding="utf-8")).items()}
    MAT = json.load(open(ortak.VERI / "matematik_oturum.json", encoding="utf-8"))
    rows = list(csv.DictReader(open(ortak.GIRDI / "tmk_cee2_sonuclari.csv", encoding="utf-8")))
    MAKS = {"fen": 14, "matematik": 27, "sosyal": 10, "ingilizce": 22, "turkce": 27}

    wb = Workbook(); ws = wb.active; ws.title = "1_Kelime_Sayısı"; ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1"); ws["A1"] = "Bölüm Kelime Sayıları — Türkçe (1.Oturum) & İngilizce (2.Oturum)"; title(ws["A1"])
    for j, h in enumerate(["Yıl", "Basamak", "Türkçe kelime", "İngilizce kelime"], 1): hdr(ws.cell(2, j, h))
    merged = {}
    for r in M:
        merged.setdefault((r["yil"], r["basamak"]), {})[r["ders"]] = r["kelime"]
    rr = 3
    for (y, b) in sorted(merged):
        d = merged[(y, b)]
        cel(ws.cell(rr, 1, y), "0"); cel(ws.cell(rr, 2, b))
        if "Türkçe" in d: cel(ws.cell(rr, 3, d["Türkçe"]), "#,##0")
        if "İngilizce" in d: cel(ws.cell(rr, 4, d["İngilizce"]), "#,##0")
        for j in range(1, 5): ws.cell(rr, j).border = BORD
        rr += 1
    for col, w in zip("ABCD", [8, 10, 14, 16]): ws.column_dimensions[col].width = w
    lastk = rr - 1

    # Yıllık özet + grafik
    ws2 = wb.create_sheet("2_Yıllık_Özet"); ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:C1"); ws2["A1"] = "Yıllık Ortalama Kelime Sayısı"; title(ws2["A1"])
    for j, h in enumerate(["Yıl", "Türkçe ort.", "İngilizce ort."], 1): hdr(ws2.cell(2, j, h))
    s1 = "'1_Kelime_Sayısı'"
    yrs_tr = {k[0] for k, v in merged.items() if "Türkçe" in v}
    yrs_en = {k[0] for k, v in merged.items() if "İngilizce" in v}
    for i, y in enumerate(range(2016, 2027)):
        r = 3 + i; cel(ws2.cell(r, 1, y), "0")
        if y in yrs_tr: ws2.cell(r, 2, f"=AVERAGEIFS({s1}!C3:C{lastk},{s1}!A3:A{lastk},A{r})"); cel(ws2.cell(r, 2), "#,##0")
        if y in yrs_en: ws2.cell(r, 3, f"=AVERAGEIFS({s1}!D3:D{lastk},{s1}!A3:A{lastk},A{r})"); cel(ws2.cell(r, 3), "#,##0")
    endr = 2 + 11
    for col, w in zip("ABC", [8, 14, 16]): ws2.column_dimensions[col].width = w
    ch = LineChart(); ch.title = "Kelime Sayısı Trendi"; ch.y_axis.title = "Kelime"; ch.x_axis.title = "Yıl"; ch.height = 9; ch.width = 17
    ch.add_data(Reference(ws2, min_col=2, max_col=3, min_row=2, max_row=endr), titles_from_data=True)
    ch.set_categories(Reference(ws2, min_col=1, min_row=3, max_row=endr)); ch.displayBlanksAs = "gap"
    ch.series[0].graphicalProperties.line.solidFill = "C0392B"; ch.series[1].graphicalProperties.line.solidFill = "2471A3"
    ws2.add_chart(ch, "E2")

    # 3 İngilizce metrikler
    ws3 = wb.create_sheet("3_İng_Metrikler"); ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:G1"); ws3["A1"] = "İngilizce — Dilbilimsel Metrikler"; title(ws3["A1"])
    for j, h in enumerate(["Yıl", "Basamak", "Kelime", "Söz/Soru", "Cümle uz.", "Kelime uz.", "MATTR"], 1): hdr(ws3.cell(2, j, h))
    rr = 3
    for r in sorted([x for x in M if x["ders"] == "İngilizce"], key=lambda x: (x["yil"], x["basamak"])):
        cel(ws3.cell(rr, 1, r["yil"]), "0"); cel(ws3.cell(rr, 2, r["basamak"]))
        cel(ws3.cell(rr, 3, r["kelime"]), "#,##0"); ws3.cell(rr, 4, f"=C{rr}/22"); cel(ws3.cell(rr, 4), "0.0")
        cel(ws3.cell(rr, 5, r["cumle_uz"]), "0.0"); cel(ws3.cell(rr, 6, r["kelime_uz"]), "0.00"); cel(ws3.cell(rr, 7, r["mattr"]), "0.000")
        for j in range(1, 8): ws3.cell(rr, j).border = BORD
        rr += 1
    for col, w in zip("ABCDEFG", [8, 9, 9, 9, 12, 12, 12]): ws3.column_dimensions[col].width = w

    # 4 Kelime çeşitliliği
    ws4 = wb.create_sheet("4_Kelime_Çeşitliliği"); ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:H1"); ws4["A1"] = "İngilizce Kelime Çeşitliliği & Nadirlik (2.Basamak)"; title(ws4["A1"])
    for j, h in enumerate(["Yıl", "Token", "Farklı", "Ort.Zipf", "İleri<4.0", "%İleri", "Nadir<3.5", "%Nadir"], 1): hdr(ws4.cell(2, j, h))
    rr = 3
    for y in sorted(VOC):
        v = VOC[y]
        cel(ws4.cell(rr, 1, y), "0"); cel(ws4.cell(rr, 2, v["token"]), "#,##0"); cel(ws4.cell(rr, 3, v["farkli"]), "#,##0")
        cel(ws4.cell(rr, 4, v["ort_zipf"]), "0.00"); cel(ws4.cell(rr, 5, v["ileri_4.0"]), "0")
        ws4.cell(rr, 6, f"=E{rr}/C{rr}"); cel(ws4.cell(rr, 6), "0.0%"); cel(ws4.cell(rr, 7, v["nadir_3.5"]), "0")
        ws4.cell(rr, 8, f"=G{rr}/C{rr}"); cel(ws4.cell(rr, 8), "0.0%")
        for j in range(1, 9): ws4.cell(rr, j).border = BORD
        rr += 1
    for col, w in zip("ABCDEFGH", [8, 9, 8, 9, 10, 8, 11, 8]): ws4.column_dimensions[col].width = w

    # 5 Okunabilirlik
    ws5 = wb.create_sheet("5_Okunabilirlik"); ws5.sheet_view.showGridLines = False
    ws5.merge_cells("A1:F1"); ws5["A1"] = "İngilizce Okunabilirlik (2.Basamak)"; title(ws5["A1"])
    for j, h in enumerate(["Yıl", "Flesch RE", "FK sınıf", "Gunning Fog", "Kelime", "Okuma dk"], 1): hdr(ws5.cell(2, j, h))
    rr = 3
    for y in sorted(RD):
        d = RD[y]
        cel(ws5.cell(rr, 1, y), "0"); cel(ws5.cell(rr, 2, d["flesch_re"]), "0.0"); cel(ws5.cell(rr, 3, d["fk_sinif"]), "0.0")
        cel(ws5.cell(rr, 4, d["fog"]), "0.0"); cel(ws5.cell(rr, 5, d["kelime"]), "#,##0")
        ws5.cell(rr, 6, f"=E{rr}/130"); cel(ws5.cell(rr, 6), "0.0")
        for j in range(1, 7): ws5.cell(rr, j).border = BORD
        rr += 1
    for col, w in zip("ABCDEF", [8, 11, 10, 12, 9, 10]): ws5.column_dimensions[col].width = w

    # 6 Sonuçlar
    ws6 = wb.create_sheet("6_CEE2_Sonuçları"); ws6.sheet_view.showGridLines = False
    ws6.merge_cells("A1:M1"); ws6["A1"] = "Lefkoşa TMK CEE-2: ham, % başarı, kohort-arındırılmış İngilizce"; title(ws6["A1"])
    subs = ["fen", "matematik", "sosyal", "ingilizce", "turkce"]; ad = {"fen": "Fen", "matematik": "Mat", "sosyal": "Sosyal", "ingilizce": "İng", "turkce": "Türkçe"}
    H = ["Yıl"] + [f"{ad[s]} ham" for s in subs] + [f"{ad[s]} %" for s in subs] + ["Diğer4 %", "İng kohort-ar"]
    for j, h in enumerate(H, 1): hdr(ws6.cell(2, j, h))
    for i, r in enumerate(rows):
        rr = 3 + i; cel(ws6.cell(rr, 1, int(r["yil"])), "0")
        for k, s in enumerate(subs): cel(ws6.cell(rr, 2 + k, float(r[s])), "0.00")
        for k, s in enumerate(subs):
            cr = get_column_letter(2 + k); ws6.cell(rr, 7 + k, f"={cr}{rr}/{MAKS[s]}"); cel(ws6.cell(rr, 7 + k), "0.0%")
        ws6.cell(rr, 12, f"=AVERAGE(G{rr},H{rr},I{rr},K{rr})"); cel(ws6.cell(rr, 12), "0.0%")
        ws6.cell(rr, 13, f"=(J{rr}-L{rr})*100"); cel(ws6.cell(rr, 13), "+0.0;-0.0")
        for j in range(1, 14): ws6.cell(rr, j).border = BORD
    for col in "ABCDEFGHIJKLM": ws6.column_dimensions[col].width = 9
    ws6.column_dimensions["L"].width = 10; ws6.column_dimensions["M"].width = 13

    # 7 Korelasyonlar
    ws7 = wb.create_sheet("7_Korelasyonlar"); ws7.sheet_view.showGridLines = False
    ws7.merge_cells("A1:F1"); ws7["A1"] = "Hangi metrik İngilizce başarısını açıklıyor?"; title(ws7["A1"])
    yy = [int(r["yil"]) for r in rows]; pct = {s: np.array([float(r[s]) / MAKS[s] * 100 for r in rows]) for s in MAKS}
    others = np.mean([pct[s] for s in ["fen", "matematik", "sosyal", "turkce"]], axis=0)
    rel = {yy[i]: pct["ingilizce"][i] - others[i] for i in range(len(yy))}
    wpq2 = {r["yil"]: r["soz_soru"] for r in M if r["ders"] == "İngilizce" and r["basamak"] == 2}
    match = sorted(set(RD) & set(rel) & set(wpq2))
    for j, h in enumerate(["Yıl", "Söz/Soru", "Okuma dk", "%İleri kelime", "Flesch RE", "Kohort-ar."], 1): hdr(ws7.cell(3, j, h))
    for i, y in enumerate(match):
        r = 4 + i
        cel(ws7.cell(r, 1, y), "0"); cel(ws7.cell(r, 2, round(wpq2[y], 1)), "0.0"); cel(ws7.cell(r, 3, RD[y]["okuma_dk"]), "0.0")
        cel(ws7.cell(r, 4, VOC[y]["ileri_4.0_oran"] / 100), "0.0%"); cel(ws7.cell(r, 5, RD[y]["flesch_re"]), "0.0")
        cel(ws7.cell(r, 6, round(rel[y], 1)), "+0.0;-0.0")
        for j in range(1, 7): ws7.cell(r, j).border = BORD
    endb = 3 + len(match)
    ws7.cell(endb + 2, 1, "KORELASYON (kohort-ar. başarı ile)").font = Font(name=FONT, bold=True, color=NAVY)
    cr0 = endb + 3
    for off, (lbl, col) in enumerate([("Söz/Soru (hacim)", "B"), ("Okuma süresi (hacim)", "C"), ("%İleri kelime (normalize)", "D"), ("Flesch (yapısal)", "E")]):
        r = cr0 + off
        ws7.cell(r, 1, lbl).font = Font(name=FONT, size=10); ws7.cell(r, 1).alignment = Alignment(horizontal="left")
        ws7.cell(r, 2, f"=CORREL({col}4:{col}{endb},F4:F{endb})"); cel(ws7.cell(r, 2), "+0.00;-0.00", bold=True)
    ws7.column_dimensions["A"].width = 28
    for col in "BCDEF": ws7.column_dimensions[col].width = 13
    bar = BarChart(); bar.type = "bar"; bar.title = "Metrik ↔ başarı korelasyonu"; bar.height = 7; bar.width = 15; bar.legend = None
    bar.add_data(Reference(ws7, min_col=2, min_row=cr0, max_row=cr0 + 3)); bar.set_categories(Reference(ws7, min_col=1, min_row=cr0, max_row=cr0 + 3))
    ws7.add_chart(bar, "H3")

    # 8 Matematik vs İng
    ws8 = wb.create_sheet("8_Matematik_vs_İng"); ws8.sheet_view.showGridLines = False
    ws8.merge_cells("A1:H1"); ws8["A1"] = "Aynı 2. Oturum: İngilizce vs Matematik yükü"; title(ws8["A1"])
    for j, h in enumerate(["Yıl", "Basamak", "İng söz/soru", "İng karakter", "Mat söz/soru", "Mat sayı", "Mat sembol", "Mat karakter"], 1): hdr(ws8.cell(2, j, h))
    rr = 3
    for r in sorted(MAT, key=lambda x: (x["yil"], x["basamak"])):
        cel(ws8.cell(rr, 1, r["yil"]), "0"); cel(ws8.cell(rr, 2, r["basamak"]))
        cel(ws8.cell(rr, 3, r["ing_soz_soru"]), "0.0"); cel(ws8.cell(rr, 4, r["ing_karakter"]), "#,##0")
        cel(ws8.cell(rr, 5, r["mat_soz_soru"]), "0.0"); cel(ws8.cell(rr, 6, r["mat_sayi"]), "0"); cel(ws8.cell(rr, 7, r["mat_sembol"]), "0"); cel(ws8.cell(rr, 8, r["mat_karakter"]), "#,##0")
        for j in range(1, 9): ws8.cell(rr, j).border = BORD
        rr += 1
    lastd = rr - 1
    for col, w in zip("ABCDEFGH", [7, 9, 12, 12, 12, 9, 10, 12]): ws8.column_dimensions[col].width = w
    base = rr + 1
    for j, h in enumerate(["Yıl", "İng söz/soru", "Mat söz/soru"], 1): hdr(ws8.cell(base, j, h))
    yset = sorted(set(r["yil"] for r in MAT))
    for i, y in enumerate(yset):
        r = base + 1 + i; cel(ws8.cell(r, 1, y), "0")
        ws8.cell(r, 2, f"=AVERAGEIFS(C3:C{lastd},A3:A{lastd},A{r})"); cel(ws8.cell(r, 2), "0.0")
        ws8.cell(r, 3, f"=AVERAGEIFS(E3:E{lastd},A3:A{lastd},A{r})"); cel(ws8.cell(r, 3), "0.0")
    yend = base + len(yset)
    ch = LineChart(); ch.title = "İngilizce şişti, Matematik şişmedi"; ch.height = 8; ch.width = 15; ch.y_axis.title = "Söz/soru"; ch.x_axis.title = "Yıl"
    ch.add_data(Reference(ws8, min_col=2, max_col=3, min_row=base, max_row=yend), titles_from_data=True)
    ch.set_categories(Reference(ws8, min_col=1, min_row=base + 1, max_row=yend))
    ch.series[0].graphicalProperties.line.solidFill = "1F6FB2"; ch.series[1].graphicalProperties.line.solidFill = "E67E22"
    ws8.add_chart(ch, "E" + str(base))

    out = ortak.CIKTI / "KGS_Ingilizce_Analiz.xlsx"
    wb.save(out)
    print(f"✓ {out.name} yazıldı (8 sekme). Excel/LibreOffice'te açınca formüller hesaplanır.")

if __name__ == "__main__":
    main()
